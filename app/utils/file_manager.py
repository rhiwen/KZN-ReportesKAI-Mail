# app/utils/file_manager.py
import os
import csv
from openpyxl import Workbook, load_workbook

os.makedirs("data", exist_ok=True)

def save_files(data, timestamp):
    csv_file = f"data/reporte_kai_{timestamp}.csv"
    xlsx_file = f"data/reporte_kai_{timestamp}.xlsx"

    # Columnas del nuevo reporte
    columns = ['TAREA', 'PROYECTO', 'TÍTULO', 'ESTADO', 'ASIGNADO A', 
               'FH CREACION', 'FH ACTUALIZACION']

    with open(csv_file, mode='w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(columns)
        for d in data:
            writer.writerow([
                d['task_id'], d['project_name'], d['title'], d['status'],
                d['assigned_to'], d['created_on'], d['updated_on']
            ])

    workbook = Workbook()
    sheet = workbook.active
    sheet.append(columns)
    
    for d in data:
        sheet.append([
            d['task_id'], d['project_name'], d['title'], d['status'],
            d['assigned_to'], d['created_on'], d['updated_on']
        ])
    
    workbook.save(xlsx_file)

    return csv_file, xlsx_file

def xlsx_to_html(xlsx_path):
    wb = load_workbook(xlsx_path)
    sheet = wb.active
    html = '<table border="1" style="border-collapse: collapse; width: 100%;">'
    
    # Estilos para la tabla HTML
    html += '<style>'
    html += 'th { background-color: #f2f2f2; padding: 8px; text-align: left; }'
    html += 'td { padding: 8px; }'
    html += 'tr:nth-child(even) { background-color: #f9f9f9; }'
    html += '</style>'
    
    # Cabeceras con <th>
    first_row = True
    for row in sheet.iter_rows(values_only=True):
        if first_row:
            html += '<tr>' + ''.join(f'<th>{cell}</th>' for cell in row) + '</tr>'
            first_row = False
        else:
            html += '<tr>' + ''.join(f'<td>{cell}</td>' for cell in row) + '</tr>'
    
    html += '</table>'
    return html